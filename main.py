

from email import encoders
from email.mime.base import MIMEBase
import openpyxl
from posixpath import basename

from PyQt5.QtWidgets import *
from PyQt5 import uic

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication


class MyGUI(QMainWindow):

    def __init__(self):
        super(MyGUI, self).__init__()
        uic.loadUi("mailGUI.ui", self)
        self.show()

        self.login.clicked.connect(self.Login)
        self.pickExcel.clicked.connect(self.PickExcel)
        self.addAttachement.clicked.connect(self.AddAttachement)
        self.send.clicked.connect(self.SendMail)

    def SendMail(self):
        dialog = QMessageBox()
        dialog.setText("Do you want to send this mail?")
        dialog.addButton(QPushButton("Yes"), QMessageBox.YesRole)
        dialog.addButton(QPushButton("No"), QMessageBox.NoRole)

        if dialog.exec_() == 0:
            try:

                self.msg["From"] = self.email.text()
                self.msg["Cc"] = self.to.text()
                self.msg["Subject"] = self.subject.text()
                body = self.mailText.toPlainText()
                self.msg.attach(MIMEText(body))
                print("passed")
                text = self.msg.as_string()
                self.server.sendmail(
                    self.email.text(), self.msg["CC"].split(","), text)

                message_box = QMessageBox()
                message_box.setText("Mail sent!")
                message_box.exec()
            except Exception as err:
                print(err)
                message_box = QMessageBox()
                message_box.setText("Sending Mail failed!")
                message_box.exec()
    msg = MIMEMultipart()

    def AddAttachement(self):
        options = QFileDialog.Options()
        filenames, _ = QFileDialog.getOpenFileNames(
            self, "Open File", "", "All Files (*.*)", options=options)
        if filenames != []:
            for filename in filenames:

                attachment = open(filename, 'rb')
                filename = filename[filename.rfind("/")+1:]
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition",
                                f"attachment; filename={filename}")
                self.msg.attach(part)
                if not self.attachements.text().endswith(":"):
                    self.attachements.setText(self.attachements.text()+",")
                self.attachements.setText(
                    self.attachements.text()+" "+basename(filename))

    def PickExcel(self):
        options = QFileDialog.Options()
        filename, _ = QFileDialog.getOpenFileName(
            self, "Open File", "", "All Files (*.*)", options=options)
        dataframe = openpyxl.load_workbook(filename)
        dataframe1 = dataframe.active
        tr = ""
        for row in range(0, dataframe1.max_row):
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                if tr != "":
                    tr += ","
                tr += col[row].value
        self.to.setText(tr)

    def Login(self):
        try:
            self.server = smtplib.SMTP(
                self.smtpServer.text(), self.port.text())
            self.server.ehlo()
            self.server.starttls()
            self.server.ehlo()
            self.server.login(self.email.text(), self.password.text())
            # disabling email,password,login btn
            self.email.setEnabled(False)
            self.password.setEnabled(False)
            self.smtpServer.setEnabled(False)
            self.port.setEnabled(False)
            self.login.setEnabled(False)

            # enaibling the other fields
            self.addAttachement.setEnabled(True)
            self.mailText.setEnabled(True)
            self.pickExcel.setEnabled(True)
            self.send.setEnabled(True)
            self.subject.setEnabled(True)
            self.to.setEnabled(True)
            self.attachements.setEnabled(True)

        except smtplib.SMTPAuthenticationError:
            message_box = QMessageBox()
            message_box.setText("Invalid Login Info!")
            message_box.exec()
        except Exception as err:
            message_box = QMessageBox()
            message_box.setText("Login failed!")
            message_box.exec()
            print(err)


app = QApplication([])
window = MyGUI()
app.exec_()
